import os
import sys
import json
from openpyxl import Workbook
import numpy as np

def interval(beg: int, end: int) -> list:
    return list(range(beg, end+1)) if beg < end else list(range(beg, end-1, -1))

version_map = dict()
version_map['cppcheck'] = [7,1,27,2,15,29,14,6,23,5,17,30,24,8,16,25,3,19,9,18,12,11,21,28,13,20,22,4,26,10]
version_map['proj'] = [24,8,13,12,11,21,22,9,18,16,1,3,17,2,4,19,10,6,20,5,27,23,14,28,15,7,26,25]
version_map['openssl'] = [13,1,2,9,12,19,14,28,11,15,5,18,16,22,10,23,26,24,25,17,6,4,7,27,20,21]
version_map['cpp_peglib'] = [8,9,7,6,5,4,2,3,1,10]
version_map['exiv2'] = [10,15,14,4,13,6,8,9,11,12,7,5,16,17,3,2,1,18,19,20]
version_map['libchewing'] = [6,8,3,4,5,7,2,1]
version_map['libxml2'] = [4,7,1,2,3,5,6]
version_map['yaml_cpp'] = [10] + list(range(1, 9+1))
version_map['thefuck'] = interval(32, 13)+interval(11, 1)
version_map['fastapi'] = interval(16, 2)
version_map['spacy'] = interval(7, 1)
version_map['youtube-dl'] = interval(35,31)+interval(10,8)+[30,7,29,28,6,27,26,5,25,24]+interval(23,18)+[3,16,2,15,1,13,12,11]

directory = os.path.abspath(os.path.dirname(os.path.dirname(sys.argv[0])))
print(directory)



def findLine(single_obj, oracle):
    ranks = list()
    for key in oracle.keys():
        lines = oracle[key]
        for line_obj in single_obj['lines']:
            if key == single_obj['files'][line_obj['index']] and line_obj['line'] in lines:
                ranks.append(line_obj['ranking'])
    return min(ranks) if len(ranks) > 0 else -1, np.mean(ranks) if len(ranks) > 0 else -1


def list2dict(l):
    ret = dict()
    for pair in l:
        ret[pair['file']] = pair['lines']
    return ret


def getTime(method: str, proj: str, index: int):
    with open(os.path.join(directory, f"log/time/{method}/{proj}/{index}.txt"), 'r') as f:
        return int(f.read().strip())


def main():
    proj = sys.argv[1].strip()
    oracle_path = sys.argv[3].strip()
    save_dir = sys.argv[4].strip()
    for method in sys.argv[2].strip().split(','):

        if proj not in version_map.keys():
            print(f"{proj} not in version_map")
            continue
        version = version_map[proj]
        
        oracle = None
        with open(oracle_path) as f:
            oracle = json.load(f)

        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(1, 1, 'version')
        sheet.cell(1, 2, 'FR')
        sheet.cell(1, 3, 'AR')
        sheet.cell(1, 4, 'test LOC')
        sheet.cell(1, 5, 'time(sec)')
        
        for idx, ver in enumerate(version):
            print(proj, method, idx+1)
            with open(os.path.join(directory, f"coverage/{method}/{proj}/{idx+1}.json"), 'r') as f:
                cov = json.load(f)
            fr, ar = findLine(cov, list2dict(oracle['version'][ver-1]))
            sheet.cell(idx + 2, 1, ver)
            sheet.cell(idx + 2, 2, fr)
            sheet.cell(idx + 2, 3, ar)
            sheet.cell(idx + 2, 4, cov['total'])
            sheet.cell(idx + 2, 5, getTime(method, proj, idx + 1))

        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        workbook.save(os.path.join(save_dir, f"{method}.xlsx"))


if __name__ == '__main__':
    main()
